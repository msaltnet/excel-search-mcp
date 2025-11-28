"""
Openpyxl Excel Adapter

For development/testing environments and regular Excel files.
Supports parallel execution, no COM dependencies.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, List

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from .adapter_base import ExcelAdapter
from .sheet_model import SheetModel, CellRange


class OpenpyxlAdapter(ExcelAdapter):
    """
    Excel adapter using openpyxl library.

    Pros:
    - Cross-platform (Windows, Mac, Linux)
    - Fast for regular Excel files
    - Supports parallel execution

    Cons:
    - Cannot read DRM-protected Excel files
    - May have formula evaluation issues
    """

    def __init__(
        self, max_rows_limit: int | None = None, max_cols_limit: int | None = None
    ) -> None:
        self.max_rows_limit = max_rows_limit
        self.max_cols_limit = max_cols_limit

    def list_sheets_from_file(self, file_path: str | Path) -> list[str]:
        """List all sheet names from Excel file."""
        path = Path(file_path)
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def get_sheet_model_from_file(
        self, file_path: str | Path, sheet_name: str
    ) -> SheetModel:
        """Extract SheetModel from Excel file."""
        path = Path(file_path)
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]

            dim = ws.calculate_dimension()
            min_col, min_row, max_col, max_row = range_boundaries(dim)

            # Optional: enforce limits
            if self.max_rows_limit and (max_row - min_row + 1) > self.max_rows_limit:
                max_row = min_row + self.max_rows_limit - 1
            if self.max_cols_limit and (max_col - min_col + 1) > self.max_cols_limit:
                max_col = min_col + self.max_cols_limit - 1

            # Read values only for range
            values_2d: List[List[Any]] = [
                list(row)
                for row in ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            ]

            # Trim trailing empty rows
            while values_2d:
                last_row = values_2d[-1]
                if all(
                    v is None or (isinstance(v, str) and v.strip() == "")
                    for v in last_row
                ):
                    values_2d.pop()
                    max_row -= 1
                else:
                    break

            # Trim trailing empty columns
            if values_2d:
                while max_col > min_col:
                    col_idx = max_col - min_col
                    if all(
                        row[col_idx] is None
                        or (isinstance(row[col_idx], str) and row[col_idx].strip() == "")
                        for row in values_2d
                    ):
                        for row in values_2d:
                            row.pop()
                        max_col -= 1
                    else:
                        break

            if not values_2d:
                return SheetModel(
                    name=sheet_name,
                    values=[],
                    used_range=CellRange(1, 1, 1, 1),
                    merged_regions=[],
                )

            used_range = CellRange(min_row, min_col, max_row, max_col)
            merged_regions = []

            return SheetModel(
                name=sheet_name,
                values=values_2d,
                used_range=used_range,
                merged_regions=merged_regions,
            )
        finally:
            wb.close()

    def shutdown(self) -> None:
        """No cleanup needed for openpyxl."""
        pass
