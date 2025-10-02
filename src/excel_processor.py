"""
Excel Processor Module

Module responsible for reading, parsing, and extracting data from Excel files
"""

import logging
from pathlib import Path
from typing import Dict, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

from .config_manager import config_manager

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Excel file processing class"""

    def __init__(self):
        self.supported_formats = config_manager.get_supported_extensions()
        self.config_manager = config_manager

    def is_supported_file(self, file_path: Path) -> bool:
        """Check if the file format is supported"""
        return file_path.suffix.lower() in self.supported_formats

    def is_file_path_within_work_directory(self, file_path: str) -> bool:
        """Check if the file path is within work directory"""
        return self.config_manager.is_path_within_work_directory(file_path)

    def validate_file_path(self, file_path: str) -> Dict[str, Any]:
        """Validate file path and return validation result"""
        try:
            path = Path(file_path)

            # Check if file exists
            if not path.exists():
                return {
                    "valid": False,
                    "error": f"File does not exist: {file_path}",
                    "error_code": "FILE_NOT_FOUND",
                }

            # Check if it's a file
            if not path.is_file():
                return {
                    "valid": False,
                    "error": f"Path is not a file: {file_path}",
                    "error_code": "NOT_A_FILE",
                }

            # Check if file is within work directory
            if not self.is_file_path_within_work_directory(file_path):
                work_dir = self.config_manager.get_work_directory()
                return {
                    "valid": False,
                    "error": f"File access denied: {file_path}. Work directory: {work_dir}",
                    "error_code": "ACCESS_DENIED",
                    "work_directory": work_dir,
                }

            # Check file size
            max_size_mb = self.config_manager.get_max_file_size_mb()
            file_size_mb = path.stat().st_size / (1024 * 1024)
            if file_size_mb > max_size_mb:
                return {
                    "valid": False,
                    "error": f"File too large: {file_size_mb:.2f}MB > {max_size_mb}MB",
                    "error_code": "FILE_TOO_LARGE",
                    "file_size_mb": file_size_mb,
                    "max_size_mb": max_size_mb,
                }

            return {
                "valid": True,
                "file_path": str(path.resolve()),
                "allowed": True,
                "file_size_mb": file_size_mb,
            }

        except Exception as e:
            return {
                "valid": False,
                "error": f"File validation error: {str(e)}",
                "error_code": "VALIDATION_ERROR",
            }

    def get_file_info(self, file_path: Path) -> Dict[str, Any]:
        """Get basic information about the Excel file"""
        try:
            # Validate file path first
            validation = self.validate_file_path(str(file_path))
            if not validation["valid"]:
                logger.warning(
                    f"File access denied: {file_path} - {validation['error']}"
                )

                return {
                    "success": False,
                    "error": validation["error"],
                    "error_code": validation["error_code"],
                    "file_path": str(file_path),
                    "work_directory": validation.get("work_directory", ""),
                }

            if not self.is_supported_file(file_path):
                return {
                    "success": False,
                    "error": f"Unsupported file format: {file_path.suffix}",
                    "supported_formats": self.supported_formats,
                }

            # Get workbook information using openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)

            worksheets = []
            for i, sheet_name in enumerate(workbook.sheetnames):
                worksheet = workbook[sheet_name]

                # Calculate worksheet size
                max_row = worksheet.max_row
                max_col = worksheet.max_column

                # Check if there is actual data
                has_data = False
                if (
                    max_row > 1 or max_col > 1
                ):  # Consider having data even if only headers exist
                    has_data = True

                worksheets.append(
                    {
                        "name": sheet_name,
                        "index": i,
                        "row_count": max_row,
                        "column_count": max_col,
                        "has_data": has_data,
                    }
                )

            workbook.close()

            # File metadata
            stat = file_path.stat()

            return {
                "success": True,
                "file_path": str(file_path.absolute()),
                "file_name": file_path.name,
                "file_size": stat.st_size,
                "worksheets": worksheets,
                "total_worksheets": len(worksheets),
                "created_time": datetime.fromtimestamp(stat.st_ctime).isoformat() + "Z",
                "modified_time": datetime.fromtimestamp(stat.st_mtime).isoformat()
                + "Z",
                "file_format": file_path.suffix.lower(),
            }

        except Exception as e:
            logger.error(f"Failed to get file information: {file_path} - {e}")
            return {
                "success": False,
                "error": f"Cannot get file information: {str(e)}",
                "file_path": str(file_path.absolute()),
            }

    def read_worksheet_data(
        self,
        file_path: Path,
        worksheet_name: Optional[str] = None,
        max_rows: Optional[int] = None,
        include_headers: bool = True,
    ) -> Dict[str, Any]:
        """Read worksheet data and convert to JSON"""
        try:
            if not self.is_supported_file(file_path):
                return {
                    "success": False,
                    "error": f"Unsupported file format: {file_path.suffix}",
                    "supported_formats": self.supported_formats,
                }

            # Read Excel file using pandas
            if worksheet_name:
                df = pd.read_excel(
                    file_path, sheet_name=worksheet_name, engine="openpyxl"
                )
            else:
                df = pd.read_excel(
                    file_path, sheet_name=0, engine="openpyxl"
                )  # First sheet

            # Limit number of rows
            if max_rows and len(df) > max_rows:
                df = df.head(max_rows)

            # Convert NaN values to None
            df = df.where(pd.notnull(df), None)

            # Convert data to dictionary
            if include_headers:
                # Use column names as headers
                headers = df.columns.tolist()
                rows = df.values.tolist()
            else:
                # Include all data with index
                headers = ["Index"] + df.columns.tolist()
                rows = []
                for idx, row in df.iterrows():
                    rows.append([idx] + row.tolist())

            # Collect data type information
            data_types = {}
            for i, col in enumerate(df.columns):
                data_types[col] = str(df[col].dtype)

            return {
                "success": True,
                "file_path": str(file_path.absolute()),
                "worksheet_name": worksheet_name or df.index.name or "Sheet1",
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
                "column_count": len(headers),
                "data_types": data_types,
                "max_rows_applied": max_rows,
                "include_headers": include_headers,
            }

        except FileNotFoundError:
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "file_path": str(file_path.absolute()),
            }
        except PermissionError:
            return {
                "success": False,
                "error": f"No permission to access file: {file_path}",
                "file_path": str(file_path.absolute()),
            }
        except Exception as e:
            logger.error(f"Failed to read worksheet data: {file_path} - {e}")
            return {
                "success": False,
                "error": f"Cannot read data: {str(e)}",
                "file_path": str(file_path.absolute()),
            }

    def get_worksheet_summary(self, file_path: Path) -> Dict[str, Any]:
        """Get summary information for all worksheets"""
        try:
            if not self.is_supported_file(file_path):
                return {
                    "success": False,
                    "error": f"Unsupported file format: {file_path.suffix}",
                    "supported_formats": self.supported_formats,
                }

            # Open workbook using openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)

            worksheets_summary = []

            for i, sheet_name in enumerate(workbook.sheetnames):
                worksheet = workbook[sheet_name]

                # Worksheet size
                max_row = worksheet.max_row
                max_col = worksheet.max_column

                # Check actual data range
                has_data = False
                data_range = None

                if max_row > 1 or max_col > 1:
                    has_data = True
                    # Find actual range with data
                    min_row = max_row
                    min_col = max_col

                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                min_row = min(min_row, cell.row)
                                min_col = min(min_col, cell.column)

                    if min_row <= max_row and min_col <= max_col:
                        data_range = {
                            "start_row": min_row,
                            "end_row": max_row,
                            "start_column": get_column_letter(min_col),
                            "end_column": get_column_letter(max_col),
                        }

                # Header information from first row (if available)
                headers = []
                if max_row > 0:
                    header_row = worksheet[1]
                    headers = [
                        cell.value for cell in header_row if cell.value is not None
                    ]

                worksheets_summary.append(
                    {
                        "name": sheet_name,
                        "index": i,
                        "row_count": max_row,
                        "column_count": max_col,
                        "has_data": has_data,
                        "data_range": data_range,
                        "headers": headers,
                        "header_count": len(headers),
                    }
                )

            workbook.close()

            return {
                "success": True,
                "file_path": str(file_path.absolute()),
                "file_name": file_path.name,
                "worksheets": worksheets_summary,
                "total_worksheets": len(worksheets_summary),
            }

        except Exception as e:
            logger.error(
                f"Failed to get worksheet summary information: {file_path} - {e}"
            )
            return {
                "success": False,
                "error": f"Cannot get worksheet information: {str(e)}",
                "file_path": str(file_path.absolute()),
            }

    def search_in_worksheet(
        self,
        file_path: Path,
        search_term: str,
        worksheet_name: Optional[str] = None,
        case_sensitive: bool = False,
    ) -> Dict[str, Any]:
        """Search for specific text in worksheet"""
        try:
            if not self.is_supported_file(file_path):
                return {
                    "success": False,
                    "error": f"Unsupported file format: {file_path.suffix}",
                    "supported_formats": self.supported_formats,
                }

            # Read Excel file using pandas
            if worksheet_name:
                df = pd.read_excel(
                    file_path, sheet_name=worksheet_name, engine="openpyxl"
                )
            else:
                df = pd.read_excel(file_path, sheet_name=0, engine="openpyxl")

            # Execute search
            if not case_sensitive:
                search_term = search_term.lower()
                search_df = df.astype(str).apply(lambda x: x.str.lower())
            else:
                search_df = df.astype(str)

            # Find search results
            matches = []
            for col_idx, col in enumerate(df.columns):
                for row_idx, value in enumerate(df[col]):
                    if case_sensitive:
                        if search_term in str(value):
                            matches.append(
                                {
                                    "row": row_idx + 1,  # 1-based indexing
                                    "column": col,
                                    "column_index": col_idx,
                                    "value": str(value),
                                    "cell_address": f"{col}{row_idx + 1}",
                                }
                            )
                    else:
                        if search_term in str(value).lower():
                            matches.append(
                                {
                                    "row": row_idx + 1,
                                    "column": col,
                                    "column_index": col_idx,
                                    "value": str(value),
                                    "cell_address": f"{col}{row_idx + 1}",
                                }
                            )

            return {
                "success": True,
                "file_path": str(file_path.absolute()),
                "worksheet_name": worksheet_name or "Sheet1",
                "search_term": search_term,
                "case_sensitive": case_sensitive,
                "total_matches": len(matches),
                "matches": matches,
            }

        except Exception as e:
            logger.error(f"Failed to search worksheet: {file_path} - {e}")
            return {
                "success": False,
                "error": f"Error occurred during search: {str(e)}",
                "file_path": str(file_path.absolute()),
            }


# Convenience functions
def get_excel_summary(file_path: str) -> Dict[str, Any]:
    """Convenience function that returns Excel file summary information"""
    processor = ExcelProcessor()
    return processor.get_file_info(Path(file_path))


def read_excel_data(
    file_path: str, worksheet_name: Optional[str] = None, max_rows: Optional[int] = None
) -> Dict[str, Any]:
    """Convenience function to read Excel file data"""
    processor = ExcelProcessor()
    return processor.read_worksheet_data(Path(file_path), worksheet_name, max_rows)


def get_worksheet_summary(file_path: str) -> Dict[str, Any]:
    """Convenience function that returns worksheet summary information"""
    processor = ExcelProcessor()
    return processor.get_worksheet_summary(Path(file_path))


def search_in_excel(
    file_path: str,
    search_term: str,
    worksheet_name: Optional[str] = None,
    case_sensitive: bool = False,
) -> Dict[str, Any]:
    """Convenience function to search text in Excel file"""
    processor = ExcelProcessor()
    return processor.search_in_worksheet(
        Path(file_path), search_term, worksheet_name, case_sensitive
    )
